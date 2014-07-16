import re
import csv
import json
import urllib2
import collections

from openpyxl import load_workbook
from openpyxl.cell import get_column_letter

from cd_data import INCOME_BY_CD
from data_districts import DATA_DISTRICTS

OUTPUT_FPATH = "./output.json"
TMP_CSV_FPATH = "output_demo_data.xlsx"
CENSUS_DATA_URL = "http://www.nyc.gov/html/dcp/download/census/census2010/t_sf1_dp_cd.xlsx"
SHEET_NAME = "CD Profiles"
SQF_FPATH = "./SQF 2012.csv"

CD_ID_MAP = {"Manhattan": 100, "Bronx": 200, "Brooklyn": 300, "Queens": 400, "StatenIsland": 500}

def extract_demographics():
    vals = collections.defaultdict(dict)

    g = urllib2.urlopen(CENSUS_DATA_URL)
    x = g.read()
    with open(TMP_CSV_FPATH, "w") as f:
        f.write(x)

    wb = load_workbook(TMP_CSV_FPATH)
    sheet = wb.get_sheet_by_name(SHEET_NAME)
    max_row = sheet.get_highest_row()
    max_col = min(sheet.get_highest_column(), 20)

    boro, dist_num = None, None

    for row_idx in xrange(1, max_row + 1):
        row = [sheet.cell('%s%d' % (get_column_letter(col_idx), row_idx)).value for col_idx in xrange(1, max_col + 1)]
        row_filtered = filter(None, row)

        if row[0]:
            m = re.match(r"(?P<boro>\w+) Community District (?P<dist_num>\d+)", row_filtered[0].replace("Staten Island", "StatenIsland"))
            if m:
                boro = m.groupdict()["boro"]
                dist_num = m.groupdict()["dist_num"]

        if boro and dist_num:
            cd_id = CD_ID_MAP[boro] + int(dist_num)

            if len(row_filtered) == 7:
                metric = row_filtered[0]
                val = row_filtered[3]
                if metric == "Total Population":
                    vals[cd_id]["total_pop"] = val
                elif metric == "White Nonhispanic":
                    vals[cd_id]["white_nonhispanic"] = val
    return vals

def load_stop_data():
    stop_data_by_precinct = dict()

    with open(SQF_FPATH, 'r') as csvfile:
        csvreader = csv.reader(csvfile)
        next(csvreader, None)
        for row in csvreader:
            precinct = int(row[1])
            arrest = bool(int(row[25]))
            frisk = bool(int(row[16]))
            search = bool(int(row[17]))
            weapon = any([bool(int(row[i])) for i in xrange(19,25)]) 
            contra = bool(int(row[18]))
            force = any([bool(int(row[i])) for i in xrange(33,42)])

            if precinct not in stop_data_by_precinct:
                stop_data_by_precinct[precinct] = collections.defaultdict(int)
            if arrest:
                stop_data_by_precinct[precinct]['arrest'] += 1
            if frisk:
                stop_data_by_precinct[precinct]['frisk'] += 1
            if search:
                stop_data_by_precinct[precinct]['search'] += 1
            if weapon:
                stop_data_by_precinct[precinct]['weapon'] += 1
            if contra:
                stop_data_by_precinct[precinct]['contra'] += 1
            if force:
                stop_data_by_precinct[precinct]['force'] += 1
            stop_data_by_precinct[precinct]['total_stops'] += 1

    return stop_data_by_precinct

def datagen():
    stop_data = load_stop_data()
    race_by_cd = extract_demographics()

    output = []
    for i, data in enumerate(DATA_DISTRICTS):

        if not any([c in race_by_cd for c in data['cds']]) or not any([p in stop_data for p in data['pcts']]):
            continue

        key = '-'.join([str(x) for x in data['pcts']])
        total_stops = sum([stop_data.get(p, {}).get('total_stops', 0) for p in data['pcts']])
        total_pop = sum([race_by_cd.get(c, {}).get('total_pop', 0) for c in data['cds']])
        pct_w_nh = sum([race_by_cd.get(c, {}).get('white_nonhispanic', 0) for c in data['cds']]) / float(total_pop)
        avg_inc = float(sum([INCOME_BY_CD.get(c, 0) * 1000. * race_by_cd.get(c, {}).get('total_pop', 0) / float(total_pop) for c in data['cds']]))
        num_stops = total_stops / (float(total_pop) / 1000)
        pct_arrest = sum([stop_data.get(p, {}).get('arrest', 0) for p in data['pcts']]) / float(total_stops)
        pct_frisk = sum([stop_data.get(p, {}).get('frisk', 0) for p in data['pcts']]) / float(total_stops)
        pct_search = sum([stop_data.get(p, {}).get('search', 0) for p in data['pcts']]) / float(total_stops)
        pct_weapon = sum([stop_data.get(p, {}).get('weapon', 0) for p in data['pcts']]) / float(total_stops)
        pct_force = sum([stop_data.get(p, {}).get('force', 0) for p in data['pcts']]) / float(total_stops)

        output.append({
            'id': 'id' + str(key),
            'race': pct_w_nh,
            'inc': avg_inc,
            'num_stops': num_stops,
            'pct_arrest': pct_arrest,
            'pct_frisk': pct_frisk,
            'pct_search': pct_search,
            'pct_weapon': pct_weapon,
            'pct_force': pct_force
       })

    with open(OUTPUT_FPATH, 'w') as f:
        f.write(json.dumps(output))

if __name__ == "__main__":
    datagen()
